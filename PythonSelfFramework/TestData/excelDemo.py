import openpyxl

Dict = {}
book = openpyxl.load_workbook("D:\\pythonProject\\pythonDemo.xlsx")

sheet = book.active
# cell = sheet.cell(row=1,column=2)
# print(cell.value)
# cell1 = sheet.cell(row=2,column=2)
# cell2 = sheet.cell(row=3,column=2)
# print(cell1.value)
# print(cell2.value)
# sheet.cell(row=2,column=2).value = "python"
# print(sheet.cell(row=2,column=2).value)

# print(sheet.max_row)
# print(sheet.max_column)
# print(sheet["C11"].value)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == "Testcase2":
        # for j in range(2, sheet.max_column+1):
        #     print(sheet.cell(row=i, column=j).value)
        for j in range(2, sheet.max_column+1):
            # Dict["lastname"] = "shetty"
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)