import openpyxl


class HomePageData:
    test_home_data = [{"firstname":"Pratik", "email":"plothe@gmail.com", "gender":"Male"},
                            {"firstname":"Abhi", "email":"Shetty@gmail.com", "gender":"Female"}]

    @staticmethod
    def get_test_data(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("D:\\pythonProject\\pytestDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                # for j in range(2, sheet.max_column+1):
                #     print(sheet.cell(row=i, column=j).value)
                for j in range(2, sheet.max_column + 1):
                    # Dict["lastname"] = "shetty"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
